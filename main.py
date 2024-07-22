from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
import time
import regex as re
from datetime import datetime
import os
from email.message import EmailMessage
import smtplib
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pytz
from dotenv import load_dotenv

def exclui_arquivo(nome_arquivo):
    # Caminho completo do arquivo (se estiver no diretório atual)
    caminho_arquivo = os.path.join(os.getcwd(), nome_arquivo)

    # Verifica se o arquivo existe antes de tentar removê-lo
    if os.path.exists(caminho_arquivo):
        os.remove(caminho_arquivo)
        print(f"O arquivo '{nome_arquivo}' foi removido com sucesso.")
    else:
        print(f"O arquivo '{nome_arquivo}' não existe no diretório atual.")

def verifica_conteudo(pesquisa):
      # Verifica se o numero de telefone tem mais de uma linha
        if re.search(r'[\r\n]+', pesquisa.get_text()): 
            retorno = "Não cadastrado" ## planilha
            return retorno
        else:
            retorno = pesquisa.get_text() ## planilha
            return retorno

#definir o fuso do brasil
brasil_fuso = pytz.timezone('America/Sao_Paulo')
# Obtém a data e hora atual
data_hora_atual = datetime.now(brasil_fuso)
# Obtém o dia atual
dia_atual = data_hora_atual.day
# Obtém o mês atual
mes_atual = str(data_hora_atual.month).zfill(2)

#link do site preço da hora
link = "https://precodahora.ba.gov.br"

servico = Service(ChromeDriverManager().install())
navegador = webdriver.Chrome(service = servico)
navegador.get(link)

#ROTINA PARA A CIDADE DE BARREIRAS:
cidade = "BARREIRAS"

#acha a barra de pesquisas e manda o nome da marca como valor
navegador.find_element('xpath', '//*[@id="fake-sbar"]').send_keys("Mitz")

#pressiona o botão de busca
navegador.find_element('xpath', '//*[@id="fake-sbar-btn"]/i').click()
time.sleep(2)

segunda_pesquisa = False
try:
    # Acha a segunda barra de pesquisa e envia "Mitz"
    navegador.find_element('xpath', '//*[@id="top-sbar"]').send_keys("Mitz")
    segunda_pesquisa = True
except Exception as e:
    # Se ocorrer uma exceção, imprime a mensagem de erro e continua o script
    segunda_pesquisa = False
    print(f"segunda barra de pesquisa não foi alçada: {e}")
    # print("Continuando o script...")

if segunda_pesquisa == True:
    #aperta a segunda "pesquisa"
    navegador.find_element('xpath', '/html/body/header/div/div[2]/div/fieldset/button[2]').click()
    time.sleep(2)

# botão da localização
navegador.find_element('xpath', '/html/body/div[2]/div/div/div[6]/button/i').click()
time.sleep(2)

#botão de usar o centro do municipio
navegador.find_element('xpath', '//*[@id="add-center"]').click()
time.sleep(2)

#barra de pesquisa da cidade
navegador.find_element('xpath', '//*[@id="modal-regions"]/div/div/div[2]/input').send_keys(cidade)
time.sleep(2)

#botão da cidade (click)
navegador.find_element('xpath', '//*[@id="sugerir-municipios"]/ul/li[2]').click()
time.sleep(2)

#botao de "aplicar" confirmar cidade
navegador.find_element('xpath', '//*[@id="aplicar"]').click()
time.sleep(2)

try:
    # Aguarde até que o modal desapareça
    WebDriverWait(navegador, 10).until(EC.invisibility_of_element_located((By.ID, "modal-loader")))

    # Agora tente clicar no botão
    navegador.find_element('xpath', '/html/body/div[2]/div/div/div[2]/button/i').click()
    time.sleep(2)
except Exception as e:
    print(f"Ocorreu um erro: {e}")

#botão para filtrar (ultimas 24horas)
# Localize o elemento (barra)
barra = navegador.find_element('xpath', '//*[@id="data"]/div/div/span/span[3]')
time.sleep(2)
# Execute JavaScript para alterar o estilo 'width' do elemento
navegador.execute_script("arguments[0].style.width = '21.4634%';", barra)
time.sleep(2)

# Localiza o icone de arrasta
icone = navegador.find_element('xpath', '//*[@id="data"]/div/div/span/span[5]')
time.sleep(2)
# Execute JavaScript para alterar o estilo 'width' do elemento
navegador.execute_script("arguments[0].style.left = '19.0244%';", icone)
time.sleep(2)

# Localiza o texto do intervalo de tempo a ser filtrado
texto = navegador.find_element('xpath','//*[@id="data"]/div/div/span/span[1]/span[6]')
time.sleep(2)
# Execute JavaScript para alterar o estilo 'width' do elemento
navegador.execute_script("arguments[0].style.left = '11.462%';", icone)
time.sleep(2)

#aplicar filtro
navegador.find_element('xpath', '//*[@id="aplicar-filtros"]').click()
time.sleep(2)

#abrir menu de filtros
navegador.find_element('xpath', '//*[@id="sort"]').click()
time.sleep(2)

#botao de "aplicar" mais recente
navegador.find_element('xpath', '//*[@id="sort"]/option[3]').click()
time.sleep(2)


html = navegador.page_source

soup = BeautifulSoup(html, 'html.parser')

hora_pesquisa = soup.find("div", class_='list-info mt-2 mb-2')
hora_texto = hora_pesquisa.find("h6").get_text()
titulo_1 = f"{hora_texto}, MUNICIPIO: {cidade}"

dados_1 = []
continua = True
num_lista = 1
num_card = 0
cont_nao_achou = 0
while continua != False:
    card_atual = soup.find("div", id=f"card_list_{num_lista}-{num_card}")
    # Verifica se o bloco do produto foi encontrado
    if card_atual:
        # Extrai as informações do produto
        numero_produto = f"Produto {num_card+1}"

        nome_produto = card_atual.find("strong").get_text()## planilha
        

        preco_produto_pesq = card_atual.find("div", style="font-size:42px;font-weight:bold; color: #000;") ## funcionamento do código
        preco_produto = verifica_conteudo(preco_produto_pesq)
        # preco_produto = preco_produto_pesq.get_text() ## planilha

        codigo_barras_pesq = preco_produto_pesq.find_next("div") ## funcionamento do código
        # Se 'codigo_barras_pesq' não é None, tentamos extrair o número
        if codigo_barras_pesq:
            match = re.search(r'\d+', codigo_barras_pesq.get_text())
            # Se 'match' não é None, então encontramos um número
            if match:
                codigo_barras = match.group()
        else:
            codigo_barras = "Não cadastrado" ## planilha

        horas_pesquisa = codigo_barras_pesq.find_next("div") ## funcionamento do código
        horas = verifica_conteudo(horas_pesquisa)
        # horas = horas_pesquisa.get_text() ## planilha

        estabelecimento_pesquisa = horas_pesquisa.find_next("div") ## funcionamento do código
        estabelecimento = verifica_conteudo(estabelecimento_pesquisa)
        # estabelecimento = estabelecimento_pesquisa.get_text() ## planilha

        localizacao_pesquisa =  estabelecimento_pesquisa.find_next("div")## funcionamento do código
        localizacao = verifica_conteudo(localizacao_pesquisa)
        # localizacao = localizacao_pesquisa.get_text() ## planilha

        dist_cidade_pesquisa = localizacao_pesquisa.find_next("div") ## funcionamento do código
        dist_cidade = verifica_conteudo(dist_cidade_pesquisa)
        # dist_cidade = dist_cidade_pesquisa.get_text() ## planilha

        contato_estabelecimento_pesquisa = dist_cidade_pesquisa.find_next("div") ## funcionamento do código
        # Verifica se o numero de telefone tem mais de uma linha
        if re.search(r'[\r\n]+', contato_estabelecimento_pesquisa.get_text()): 
            contato_estabelecimento = "Sem telefone cadastrado" ## planilha
        else:
            contato_estabelecimento = contato_estabelecimento_pesquisa.get_text() ## planilha
        
        # Adiciona os dados à lista
        dados_1.append([numero_produto, nome_produto, preco_produto, codigo_barras, horas, estabelecimento, localizacao, dist_cidade, contato_estabelecimento])

        num_card += 1

    else:
        # print(f"Bloco do produto card numero {num_card}, não encontrado. Na lista {num_lista}")
        num_lista += 1
        cont_nao_achou += 1
        num_card = 1  # Reinicia o contador de card após percorrer toda a lista
        if cont_nao_achou == 3:
            continua =False

# Cria um DataFrame a partir dos dados
df_1 = pd.DataFrame(dados_1, columns=['Número Produto','Nome do Produto', 'Preço do Produto (R$)', 'Código de Barras', 'Horas', 'Estabelecimento', 'Localização', 'Distância até o Município', 'Contato do Estabelecimento'])

# Remove o símbolo "R$" e a vírgula dos valores da coluna 'Preço do Produto' e converte para float
df_1['Preço do Produto (R$)'] = df_1['Preço do Produto (R$)'].str.replace('R\$|\.', '', regex=True).str.replace(',', '.', regex=False).astype(float)
# Calcula o total dos preços dos produtos
total_vendido = df_1['Preço do Produto (R$)'].sum()

# Adiciona uma linha ao final do DataFrame com o valor total vendido
linha_total = pd.DataFrame({'Número Produto': ['Total Vendido'], 'Nome do Produto': [''], 'Preço do Produto (R$)': [total_vendido], 'Código de Barras': [''], 'Horas': [''], 'Estabelecimento': [''], 'Localização': [''], 'Distância até o Município': [''], 'Contato do Estabelecimento': ['']})

# Concatena o DataFrame original com a linha do total vendido
df_1 = pd.concat([df_1, linha_total], ignore_index=True)

#variavel que apenas guarda o nome da planilha
planilha_1 = f'vendas_{cidade}_({dia_atual}_{mes_atual}).xlsx'

# Salva o DataFrame em um arquivo Excel
df_1.to_excel(planilha_1, index=False)

# Ajusta a largura das colunas
wb = load_workbook(planilha_1)
ws = wb.active

# Define a largura das colunas automaticamente
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter  # Obtém a letra da coluna
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column_letter].width = adjusted_width

wb.save(planilha_1)

time.sleep(5)

##########################################################################################################################################################################################
##########################################################################################################################################################################################

#ROTINA PARA o proximo municipio: LUIS EDUARDO DE MAGALHAES
cidade = "LUÍS EDUARDO MAGALHÃES"

# botão de selecionar localização
navegador.find_element('xpath', '/html/body/div[2]/div/div/div[6]/button/i').click()
time.sleep(2)

# botão de selecionar municipio
navegador.find_element('xpath', '//*[@id="add-center"]').click()
time.sleep(2)

# escrever municipio
navegador.find_element('xpath', '//*[@id="modal-regions"]/div/div/div[2]/input').send_keys(cidade)
time.sleep(3)

# selecionar municipio
navegador.find_element('xpath', '//*[@id="sugerir-municipios"]/ul/li[2]').click()
time.sleep(2)

# aplicar municipio
navegador.find_element('xpath', '//*[@id="aplicar"]').click()
time.sleep(2)

try:
    # Aguarde até que o modal desapareça
    WebDriverWait(navegador, 10).until(EC.invisibility_of_element_located((By.ID, "modal-loader")))

    # Agora tente clicar no botão
    navegador.find_element('xpath', '/html/body/div[2]/div/div/div[2]/button/i').click()
except Exception as e:
    print(f"Ocorreu um erro: {e}")

#####

#botão para filtrar (ultimas 24horas)
# Localize o elemento (barra)
barra = navegador.find_element('xpath', '//*[@id="data"]/div/div/span/span[3]')
time.sleep(2)
# Execute JavaScript para alterar o estilo 'width' do elemento
navegador.execute_script("arguments[0].style.width = '21.4634%';", barra)
time.sleep(2)

# Localiza o icone de arrasta
icone = navegador.find_element('xpath', '//*[@id="data"]/div/div/span/span[5]')
time.sleep(2)
# Execute JavaScript para alterar o estilo 'width' do elemento
navegador.execute_script("arguments[0].style.left = '19.0244%';", icone)
time.sleep(2)

# Localiza o texto do intervalo de tempo a ser filtrado
texto = navegador.find_element('xpath','//*[@id="data"]/div/div/span/span[1]/span[6]')
time.sleep(2)
# Execute JavaScript para alterar o estilo 'width' do elemento
navegador.execute_script("arguments[0].style.left = '11.462%';", icone)
time.sleep(2)

#aplicar filtro
navegador.find_element('xpath', '//*[@id="aplicar-filtros"]').click()
time.sleep(2)

#abrir menu de filtros
navegador.find_element('xpath', '//*[@id="sort"]').click()
time.sleep(2)

#botao de "aplicar" mais recente
navegador.find_element('xpath', '//*[@id="sort"]/option[3]').click()
time.sleep(2)

#####

html = navegador.page_source

soup = BeautifulSoup(html, 'html.parser')

dados_2 =[] # lista onde serão armazenados os dados de cada linha da tabela
hora_pesquisa = soup.find("div", class_='list-info mt-2 mb-2') # pesquisa que resultara no titulo da planilha
hora_texto = hora_pesquisa.find("h6").get_text()
titulo_2 = f"{hora_texto}, MUNICIPIO: {cidade}" ## titulo planilha

continua = True
num_lista = 1
num_card = 0
cont_nao_achou = 0

while continua != False:
    card_atual = soup.find("div", id=f"card_list_{num_lista}-{num_card}")
    # Verifica se o bloco do produto foi encontrado
    if card_atual:
        # Extrai as informações do produto
        numero_produto = f"Produto {num_card+1}"

        nome_produto = card_atual.find("strong").get_text() ## planilha

        preco_produto_pesq = card_atual.find("div", style="font-size:42px;font-weight:bold; color: #000;") ## funcionamento do código
        preco_produto = verifica_conteudo(preco_produto_pesq)
        # preco_produto = preco_produto_pesq.get_text() ## planilha

        codigo_barras_pesq = preco_produto_pesq.find_next("div") ## funcionamento do código
        # Se 'codigo_barras_pesq' não é None, tentamos extrair o número
        if codigo_barras_pesq:
            match = re.search(r'\d+', codigo_barras_pesq.get_text())
            # Se 'match' não é None, então encontramos um número
            if match:
                codigo_barras = match.group()
        else:
            codigo_barras = "Não cadastrado" ## planilha

        horas_pesquisa = codigo_barras_pesq.find_next("div") ## funcionamento do código
        horas = verifica_conteudo(horas_pesquisa)
        # horas = horas_pesquisa.get_text() ## planilha

        estabelecimento_pesquisa = horas_pesquisa.find_next("div") ## funcionamento do código
        estabelecimento = verifica_conteudo(estabelecimento_pesquisa)
        # estabelecimento = estabelecimento_pesquisa.get_text() ## planilha

        localizacao_pesquisa =  estabelecimento_pesquisa.find_next("div")## funcionamento do código
        localizacao = verifica_conteudo(localizacao_pesquisa)
        # localizacao = localizacao_pesquisa.get_text() ## planilha

        dist_cidade_pesquisa = localizacao_pesquisa.find_next("div") ## funcionamento do código
        dist_cidade = verifica_conteudo(dist_cidade_pesquisa)
        # dist_cidade = dist_cidade_pesquisa.get_text() ## planilha

        contato_estabelecimento_pesquisa = dist_cidade_pesquisa.find_next("div") ## funcionamento do código
        # Verifica se o numero de telefone tem mais de uma linha
        if re.search(r'[\r\n]+', contato_estabelecimento_pesquisa.get_text()): 
            contato_estabelecimento = "Sem telefone cadastrado" ## planilha
        else:
            contato_estabelecimento = contato_estabelecimento_pesquisa.get_text() ## planilha
        
        # Adiciona os dados à lista
        dados_2.append([numero_produto, nome_produto, preco_produto, codigo_barras, horas, estabelecimento, localizacao, dist_cidade, contato_estabelecimento])

        num_card += 1
    else:
        # print(f"Bloco do produto card numero {num_card}, não encontrado. Na lista {num_lista}")
        num_lista += 1
        cont_nao_achou += 1
        num_card = 1  # Reinicia o contador de card após percorrer toda a lista
        if cont_nao_achou == 3:
            continua =False

# Cria um DataFrame a partir dos dados
df_2 = pd.DataFrame(dados_2, columns=['Número Produto','Nome do Produto', 'Preço do Produto (R$)', 'Código de Barras', 'Horas', 'Estabelecimento', 'Localização', 'Distância até o Município', 'Contato do Estabelecimento'])

# Remove o símbolo "R$" e a vírgula dos valores da coluna 'Preço do Produto' e converte para float
df_2['Preço do Produto (R$)'] = df_2['Preço do Produto (R$)'].str.replace('R\$|\.', '', regex=True).str.replace(',', '.', regex=False).astype(float)
# Calcula o total dos preços dos produtos
total_vendido = df_2['Preço do Produto (R$)'].sum()

# Adiciona uma linha ao final do DataFrame com o valor total vendido
linha_total = pd.DataFrame({'Número Produto': ['Total Vendido'], 'Nome do Produto': [''], 'Preço do Produto (R$)': [total_vendido], 'Código de Barras': [''], 'Horas': [''], 'Estabelecimento': [''], 'Localização': [''], 'Distância até o Município': [''], 'Contato do Estabelecimento': ['']})

# Concatena o DataFrame original com a linha do total vendido
df_2 = pd.concat([df_2, linha_total], ignore_index=True)

#variavel que apenas guarda o nome da planilha
planilha_2 = f'vendas_{cidade}_({dia_atual}_{mes_atual}).xlsx'

# Salva o DataFrame em um arquivo Excel
df_2.to_excel(planilha_2, index=False)

# Ajusta a largura das colunas
wb = load_workbook(planilha_2)
ws = wb.active

# Define a largura das colunas automaticamente
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter  # Obtém a letra da coluna
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column_letter].width = adjusted_width

wb.save(planilha_2)

############################################################
# Carregar variáveis de ambiente do arquivo .env
load_dotenv()
#Rotina para enviar o email
EMAIL = os.getenv("EMAIL_SMPT")
SENHA = os.getenv("SENHA_SMPT")

msg = EmailMessage()
msg['Subject'] = f'Planilhas do dia {dia_atual}/{mes_atual}'
msg['From'] = EMAIL
# msg['To'] = f'{os.getenv("EMAIL_DEV")}, {os.getenv("EMAIL_EIZEMITZ")}'
msg['To'] = os.getenv("EMAIL_DEV")

# Corpo do e-mail (texto)
corpo_email = f"""
Olá,

Segue anexo duas planilhas do dia {dia_atual}/{mes_atual}.

//{titulo_1}

//{titulo_2}

"""

msg.set_content(corpo_email)

# Anexa os arquivos (planilhas)
with open(planilha_1, 'rb') as file1:
    msg.add_attachment(file1.read(), maintype='application', subtype='octet-stream', filename=planilha_1)

with open(planilha_2, 'rb') as file2:
    msg.add_attachment(file2.read(), maintype='application', subtype='octet-stream', filename=planilha_2)


with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
    smtp.login(EMAIL, SENHA)
    smtp.send_message(msg)

print("email enviado")

# Aguarde alguns segundos para garantir que a página seja carregada
time.sleep(50)

exclui_arquivo(planilha_1)
exclui_arquivo(planilha_2)

# Feche o navegador após a conclusão
navegador.quit()